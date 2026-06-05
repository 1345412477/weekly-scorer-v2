"""周报 API"""
import os
import uuid
import shutil
from io import BytesIO
from datetime import date, datetime, timedelta
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc

from pydantic import BaseModel
from app.database import get_db
from app.models.models import WeeklyReport, ReportScore, Person
from app.schemas.schemas import ReportCreate, ReportResponse


class BatchDeleteRequest(BaseModel):
    report_ids: list[str]


class BatchDeleteResponse(BaseModel):
    message: str
    deleted_count: int
    deleted_ids: list[str]
from app.services.scoring import trigger_scoring
from app.services.ai_scorer import AIScoringError
from app.services.document_parser import (
    parse_report, extract_week_dates, get_template_path,
    classify_report_week, get_current_week, SUPPORTED_EXTENSIONS,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/api/v1/reports", tags=["周报管理"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def iso_week_number(d):
    if isinstance(d, str):
        d = date.fromisoformat(d)
    return d.isocalendar()[1]


def build_report_dict(r, scores=None):
    ws = r.week_start
    we = r.week_end
    ws_str = ws.isoformat() if hasattr(ws, "isoformat") else str(ws)
    we_str = we.isoformat() if hasattr(we, "isoformat") else str(we)
    return {
        "id": r.id,
        "author_name": r.author_name,
        "department": r.department or "",
        "person_id": r.person_id,
        "department_id": r.department_id,
        "week_start": ws_str,
        "week_end": we_str,
        "week_num": iso_week_number(ws),
        "content": r.content,
        "status": r.status,
        "report_type": getattr(r, "report_type", "normal") or "normal",
        "week_diff": getattr(r, "week_diff", 0) or 0,
        "total_score": float(scores.total_score) if scores else None,
        "grade": scores.grade if scores else None,
        "original_filename": r.original_filename,
        "submit_time": r.submit_time.isoformat() if r.submit_time else None,
        "score_time": r.score_time.isoformat() if r.score_time else None,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


@router.get("/template/download")
async def download_template():
    """下载 Excel 周报模板"""
    template_path = get_template_path()
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="模板文件不存在")

    monday, sunday = get_current_week()
    filename = f"周报模板_{monday.strftime('%Y%m%d')}-{sunday.strftime('%Y%m%d')}.xlsx"

    return FileResponse(
        path=template_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/upload")
async def upload_report(
    file: UploadFile = File(...),
    person_id: Optional[str] = Form(None),
    department_id: Optional[str] = Form(None),
    author_name: str = Form("匿名"),
    department: str = Form(""),
    confirmed_week_start: Optional[str] = Form(None),
    confirmed_week_end: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """上传周报文件，支持 .xlsx/.xls/.docx/.pdf"""
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件格式: {file_ext}，支持：{', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    if person_id:
        person_r = await db.execute(select(Person).where(Person.id == person_id))
        person = person_r.scalar_one_or_none()
        if person:
            author_name = person.name
            if not department and person.department_name:
                department = person.department_name
            if not department_id and person.department_id:
                department_id = person.department_id

    file_id = str(uuid.uuid4())
    saved_filename = f"{file_id}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, saved_filename)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        parsed = parse_report(file_path)
        week_start, week_end = extract_week_dates(file_path)

        if confirmed_week_start and confirmed_week_end:
            week_start = date.fromisoformat(confirmed_week_start)
            week_end = date.fromisoformat(confirmed_week_end)

        classification = classify_report_week(week_start, week_end)

        if classification["is_future"]:
            os.remove(file_path)
            raise HTTPException(status_code=400, detail=classification["message"])

        content = parsed["raw_content"]

        if not content or len(content.strip()) < 20:
            os.remove(file_path)
            raise HTTPException(status_code=400, detail="文件内容为空或格式不正确，无法解析周报内容")

        if week_start is None:
            monday, sunday = get_current_week()
            week_start = monday
            week_end = sunday

        report_type = classification["report_type"]
        week_diff = classification["week_diff"]

        report = WeeklyReport(
            id=file_id,
            author_name=author_name,
            department=department,
            person_id=person_id,
            department_id=department_id,
            week_start=week_start,
            week_end=week_end,
            content=content,
            file_path=file_path,
            original_filename=file.filename,
            status="submitted",
            report_type=report_type,
            week_diff=week_diff,
            submit_time=datetime.utcnow(),
        )
        db.add(report)
        await db.commit()
        await db.refresh(report)

        score_result = None
        scoring_error = None
        try:
            score_result = await trigger_scoring(report.id, db)
        except AIScoringError as e:
            scoring_error = str(e)
        except Exception as e:
            scoring_error = f"评分失败：{str(e)}"

        result = {
            "message": "上传成功",
            "report_id": report.id,
            "report_type": report_type,
            "week_diff": week_diff,
            "classification_message": classification["message"],
            "needs_confirmation": classification["needs_confirmation"],
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "content_preview": content[:300] + "..." if len(content) > 300 else content,
            "total_score": score_result["total_score"] if score_result else None,
            "grade": score_result["grade"] if score_result else None,
            "scoring_error": scoring_error,
        }

        if scoring_error:
            result["message"] = f"上传成功，但评分失败：{scoring_error}"
        elif report_type == "catch_up":
            result["message"] = f"补周报上传成功（{week_diff}周前）"
        elif report_type == "normal":
            result["message"] = "本周周报上传成功"

        if classification["needs_confirmation"]:
            if scoring_error:
                result["message"] = f"上传成功，但未能自动识别周报时间，且评分失败：{scoring_error}"
            else:
                result["message"] = "上传成功，但未能自动识别周报时间，请确认"

        return result

    except HTTPException:
        raise
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"文件解析失败: {str(e)}")


@router.post("")
async def create_report(req: ReportCreate, db: AsyncSession = Depends(get_db)):
    """创建/保存周报草稿"""
    monday, sunday = get_current_week()
    ws = req.week_start or monday
    we = req.week_end or sunday

    report = WeeklyReport(
        id=str(uuid.uuid4()),
        author_name=req.author_name,
        department=req.department or "",
        person_id=req.person_id,
        department_id=req.department_id,
        week_start=ws,
        week_end=we,
        content=req.content,
        template_id=req.template_id,
        status="draft",
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)

    return {"message": "周报已保存", "id": report.id, "status": "draft"}


@router.get("")
async def list_reports(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    author_name: Optional[str] = None,
    department: Optional[str] = None,
    is_catch_up: Optional[str] = None,
    sort_by: Optional[str] = Query(None, description="排序字段：week_num, submit_time, total_score"),
    sort_order: Optional[str] = Query("desc", description="排序方向：asc, desc"),
    db: AsyncSession = Depends(get_db),
):
    """获取周报列表，支持筛选和排序"""
    query = select(WeeklyReport)
    count_query = select(func.count()).select_from(WeeklyReport)

    # 状态筛选
    if status:
        statuses = status.split(",")
        query = query.where(WeeklyReport.status.in_(statuses))
        count_query = count_query.where(WeeklyReport.status.in_(statuses))

    # 提交人筛选
    if author_name:
        query = query.where(WeeklyReport.author_name.ilike(f"%{author_name}%"))
        count_query = count_query.where(WeeklyReport.author_name.ilike(f"%{author_name}%"))

    # 部门筛选
    if department:
        query = query.where(WeeklyReport.department.ilike(f"%{department}%"))
        count_query = count_query.where(WeeklyReport.department.ilike(f"%{department}%"))

    # 是否补周报筛选
    if is_catch_up:
        if is_catch_up == "yes":
            query = query.where(WeeklyReport.report_type == "catch_up")
            count_query = count_query.where(WeeklyReport.report_type == "catch_up")
        elif is_catch_up == "no":
            query = query.where(WeeklyReport.report_type != "catch_up")
            count_query = count_query.where(WeeklyReport.report_type != "catch_up")

    # 排序
    sort_column = WeeklyReport.created_at  # 默认排序
    if sort_by == "week_num":
        sort_column = WeeklyReport.week_start
    elif sort_by == "submit_time":
        sort_column = WeeklyReport.submit_time
    elif sort_by == "total_score":
        # 需要关联 ReportScore 表进行排序
        query = query.outerjoin(ReportScore, WeeklyReport.id == ReportScore.report_id)
        count_query = count_query.outerjoin(ReportScore, WeeklyReport.id == ReportScore.report_id)
        sort_column = ReportScore.total_score

    if sort_order == "asc":
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())

    total_r = await db.execute(count_query)
    total = total_r.scalar() or 0

    result = await db.execute(query.offset((page - 1) * size).limit(size))
    reports = result.scalars().all()

    report_ids = [r.id for r in reports]
    scores_map = {}
    if report_ids:
        scores_r = await db.execute(
            select(ReportScore).where(ReportScore.report_id.in_(report_ids))
        )
        for s in scores_r.scalars().all():
            scores_map[s.report_id] = s

    items = [build_report_dict(r, scores_map.get(r.id)) for r in reports]

    return {"items": items, "total": total, "page": page, "size": size}


@router.get("/{report_id}")
async def get_report(report_id: str, db: AsyncSession = Depends(get_db)):
    """获取周报详情（含评分维度）"""
    result = await db.execute(
        select(WeeklyReport).where(WeeklyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="周报不存在")

    scores_r = await db.execute(
        select(ReportScore).where(ReportScore.report_id == report_id)
    )
    scores = scores_r.scalar_one_or_none()

    ws = report.week_start
    we = report.week_end
    ws_str = ws.isoformat() if hasattr(ws, "isoformat") else str(ws)
    we_str = we.isoformat() if hasattr(we, "isoformat") else str(we)

    dim_scores = []
    if scores and scores.dimension_scores:
        dim_scores = scores.dimension_scores

    return {
        "id": report.id,
        "author_name": report.author_name,
        "department": report.department or "",
        "person_id": report.person_id,
        "department_id": report.department_id,
        "week_start": ws_str,
        "week_end": we_str,
        "week_num": iso_week_number(ws),
        "content": report.content,
        "status": report.status,
        "report_type": getattr(report, "report_type", "normal") or "normal",
        "week_diff": getattr(report, "week_diff", 0) or 0,
        "total_score": float(scores.total_score) if scores else None,
        "grade": scores.grade if scores else None,
        "dimension_scores": dim_scores,
        "ai_comment": scores.ai_comment if scores else None,
        "ai_suggestion": scores.ai_suggestion if scores else None,
        "original_filename": report.original_filename,
        "submit_time": report.submit_time.isoformat() if report.submit_time else None,
        "score_time": report.score_time.isoformat() if report.score_time else None,
        "created_at": report.created_at.isoformat() if report.created_at else None,
    }


@router.post("/{report_id}/submit")
async def submit_report(report_id: str, db: AsyncSession = Depends(get_db)):
    """提交周报并触发评分"""
    result = await db.execute(
        select(WeeklyReport).where(WeeklyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="周报不存在")
    if report.status != "draft":
        raise HTTPException(status_code=400, detail="周报已提交，不可重复提交")

    report.status = "submitted"
    report.submit_time = datetime.utcnow()
    await db.commit()

    try:
        score_result = await trigger_scoring(report_id, db)
        return {
            "message": "提交并评分成功",
            "report_id": report_id,
            "total_score": score_result["total_score"],
            "grade": score_result["grade"],
        }
    except AIScoringError as e:
        return {
            "message": f"提交成功，但评分失败：{str(e)}",
            "report_id": report_id,
            "total_score": None,
            "grade": None,
            "scoring_error": str(e),
        }
    except Exception as e:
        error_msg = f"评分失败：{str(e)}"
        return {
            "message": f"提交成功，但{error_msg}",
            "report_id": report_id,
            "total_score": None,
            "grade": None,
            "scoring_error": error_msg,
        }


@router.get("/{report_id}/download")
async def download_report(report_id: str, db: AsyncSession = Depends(get_db)):
    """下载周报文件"""
    result = await db.execute(
        select(WeeklyReport).where(WeeklyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="周报不存在")

    if not report.file_path or not os.path.exists(report.file_path):
        raise HTTPException(status_code=404, detail="周报文件不存在或已被删除")

    filename = report.original_filename or f"周报_{report_id}.xlsx"

    # 根据文件扩展名确定 MIME 类型
    file_ext = os.path.splitext(filename)[1].lower()
    mime_types = {
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.pdf': 'application/pdf',
    }
    media_type = mime_types.get(file_ext, 'application/octet-stream')

    return FileResponse(
        path=report.file_path,
        filename=filename,
        media_type=media_type,
    )


@router.delete("/batch", response_model=BatchDeleteResponse)
async def batch_delete_reports(
    req: BatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
):
    """批量删除周报"""
    report_ids = req.report_ids
    if not report_ids or len(report_ids) == 0:
        raise HTTPException(status_code=400, detail="请选择要删除的周报")

    if len(report_ids) > 100:
        raise HTTPException(status_code=400, detail="单次最多删除100条记录")

    result = await db.execute(
        select(WeeklyReport).where(WeeklyReport.id.in_(report_ids))
    )
    reports = result.scalars().all()

    deleted_count = 0
    deleted_ids = []
    for report in reports:
        # 删除关联的评分记录
        scores_r = await db.execute(
            select(ReportScore).where(ReportScore.report_id == report.id)
        )
        score = scores_r.scalar_one_or_none()
        if score:
            await db.delete(score)

        # 删除文件
        if report.file_path and os.path.exists(report.file_path):
            os.remove(report.file_path)

        await db.delete(report)
        deleted_count += 1
        deleted_ids.append(report.id)

    await db.commit()

    # 记录操作日志（实际项目中应存储到数据库）
    print(f"[批量删除] 操作时间: {datetime.utcnow()}, 删除数量: {deleted_count}, 删除ID: {deleted_ids}")

    return {
        "message": f"成功删除{deleted_count}条周报",
        "deleted_count": deleted_count,
        "deleted_ids": deleted_ids,
    }


@router.delete("/{report_id}")
async def delete_report(report_id: str, db: AsyncSession = Depends(get_db)):
    """删除周报"""
    result = await db.execute(
        select(WeeklyReport).where(WeeklyReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="周报不存在")

    # 删除关联的评分记录
    scores_r = await db.execute(
        select(ReportScore).where(ReportScore.report_id == report_id)
    )
    score = scores_r.scalar_one_or_none()
    if score:
        await db.delete(score)

    # 删除文件
    if report.file_path and os.path.exists(report.file_path):
        os.remove(report.file_path)

    await db.delete(report)
    await db.commit()
    return {"message": "周报已删除"}


@router.get("/export")
async def export_reports(
    status: Optional[str] = None,
    author_name: Optional[str] = None,
    department: Optional[str] = None,
    is_catch_up: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """导出周报列表为Excel文件"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="导出功能不可用，缺少openpyxl依赖")

    query = select(WeeklyReport)

    # 状态筛选
    if status:
        statuses = status.split(",")
        query = query.where(WeeklyReport.status.in_(statuses))

    # 提交人筛选
    if author_name:
        query = query.where(WeeklyReport.author_name.ilike(f"%{author_name}%"))

    # 部门筛选
    if department:
        query = query.where(WeeklyReport.department.ilike(f"%{department}%"))

    # 是否补周报筛选
    if is_catch_up:
        if is_catch_up == "yes":
            query = query.where(WeeklyReport.report_type == "catch_up")
        elif is_catch_up == "no":
            query = query.where(WeeklyReport.report_type != "catch_up")

    query = query.order_by(desc(WeeklyReport.created_at))
    query = query.limit(1000)  # 限制导出数量

    result = await db.execute(query)
    reports = result.scalars().all()

    report_ids = [r.id for r in reports]
    scores_map = {}
    if report_ids:
        scores_r = await db.execute(
            select(ReportScore).where(ReportScore.report_id.in_(report_ids))
        )
        for s in scores_r.scalars().all():
            scores_map[s.report_id] = s

    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "周报列表"

    # 设置表头
    headers = [
        "序号", "周次", "提交人", "部门", "评分", "等级",
        "是否补周报", "提交时间", "状态", "原始文件名"
    ]

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 填充数据
    row_num = 2
    for report in reports:
        scores = scores_map.get(report.id)
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=f"第{iso_week_number(report.week_start)}周").border = thin_border
        ws.cell(row=row_num, column=3, value=report.author_name).border = thin_border
        ws.cell(row=row_num, column=4, value=report.department or "-").border = thin_border
        ws.cell(row=row_num, column=5, value=float(scores.total_score) if scores else "-").border = thin_border
        ws.cell(row=row_num, column=6, value=scores.grade if scores else "-").border = thin_border
        ws.cell(row=row_num, column=7, value="是" if report.report_type == "catch_up" else "否").border = thin_border
        
        # 格式化时间为北京时间
        submit_time = ""
        if report.submit_time:
            submit_time = report.submit_time.strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row_num, column=8, value=submit_time).border = thin_border
        
        ws.cell(row=row_num, column=9, value={"draft": "草稿", "submitted": "已提交", "scored": "已评分"}.get(report.status, report.status)).border = thin_border
        ws.cell(row=row_num, column=10, value=report.original_filename or "-").border = thin_border
        row_num += 1

    # 设置列宽
    column_widths = [8, 12, 15, 20, 8, 8, 12, 22, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    filename = f"周报列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )
