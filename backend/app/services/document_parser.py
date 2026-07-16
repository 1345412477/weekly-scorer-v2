"""文档解析服务 - 支持 Excel / Word / PDF 周报解析"""
import os
import re
import logging
from datetime import date, timedelta
from typing import Optional

from app.utils.time_utils import bj_today

import openpyxl

logger = logging.getLogger(__name__)


from app.config import get_settings

# 模板文件放在 backend/templates/ 下，由 config.TEMPLATE_DIR 配置路径
# 默认 "./templates"，相对于 backend/ 目录
_template_rel = get_settings().TEMPLATE_DIR.lstrip("./\\")
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    _template_rel,
    "周报模板.xlsx",
)


def get_template_path() -> str:
    return TEMPLATE_PATH


SUPPORTED_EXTENSIONS = {".xlsx"}


# ── 文件名解析与提交人识别 ──

def extract_author_from_filename(original_filename: str) -> tuple[str | None, str]:
    """
    从文件名解析提交人姓名。

    规范格式（严格要求）：
      「提交人名字-YYYY年MM月第N周周报YYYYMMDD.xlsx」
      例如：张三-2026年6月第2周周报20260614.xlsx

    逻辑：
      1. 去掉 .xlsx 后缀
      2. 按英文「-」或全角「—」分割，取首段
      3. 首段必须为 2-15 位中文姓名（仅允许中文）
      4. 其余段不做严格校验（仅作为周次信息，不影响识别）

    返回：
      (candidate_name, hint_msg)
      - candidate_name: 识别到的中文姓名（未命中 persons 表时也返回）
      - hint_msg: 空字符串表示成功识别；否则为失败原因的中文提示
    """
    if not original_filename or not isinstance(original_filename, str):
        return None, "文件名为空"

    stem = os.path.splitext(original_filename)[0].strip()
    # 兼容全角「—」
    stem = stem.replace("—", "-")

    if "-" not in stem:
        return None, "文件名不符合规范（缺少「-」分隔符）"

    parts = [p for p in stem.split("-") if p.strip()]
    if not parts:
        return None, "文件名无法解析"

    candidate = parts[0].strip()

    # 严格要求：中文姓名 2-15 位
    import re as _re
    if not _re.fullmatch(r"[\u4e00-\u9fa5]{2,15}", candidate):
        return None, "文件名首段必须为中文姓名（2-15 位）"

    return candidate, ""


# ── 元数据扫描（汇报人 / 部门 / 日期等）──

_META_LABELS: dict[str, list[str]] = {
    "reporter": ["汇报人", "提交人", "报告人", "姓名", "填报人"],
    "department": ["部门", "所属部门", "所在部门", "单位"],
    "week_dates": ["周报周期", "周期", "日期", "时间", "报告日期"],
}


def _safe_path(file_path: str) -> bool:
    """拒绝路径遍历：不允许 .. 或绝对路径指向上传目录之外。"""
    if not file_path or not isinstance(file_path, str):
        return False
    abs_path = os.path.abspath(file_path)
    upload_root = os.path.abspath(
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
    )
    # 允许读取模板路径（TEMPLATE_PATH）及 uploads 目录
    template_root = os.path.abspath(os.path.dirname(TEMPLATE_PATH))
    return abs_path.startswith(upload_root) or abs_path.startswith(template_root)


def _safe_load_workbook(file_path: str):
    """安全加载 Excel 文件，自动释放句柄。"""
    with open(file_path, "rb") as f:
        return openpyxl.load_workbook(f, data_only=True)


def parse_report(file_path: str) -> dict:
    if not _safe_path(file_path):
        raise ValueError(f"文件路径不安全或为空: {file_path}")
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx")
    return _parse_excel(file_path)


def extract_week_dates(file_path: str) -> tuple[Optional[date], Optional[date]]:
    if not _safe_path(file_path):
        return None, None
    ext = os.path.splitext(file_path)[1].lower()
    if ext in SUPPORTED_EXTENSIONS:
        # 优先从文件内容中提取
        d1, d2 = _extract_dates_from_excel(file_path)
        if d1 and d2:
            return d1, d2
        # 内容提取失败时，尝试从文件名中提取日期
        return _extract_dates_from_filename(file_path)
    return None, None


def _extract_dates_from_filename(file_path: str) -> tuple[Optional[date], Optional[date]]:
    """从文件名中提取日期，作为备用方案。
    支持格式：
      - 张三-2026年6月第2周周报20260614.xlsx  → 从 YYYYMMDD 提取
      - 张三-2026年7月第3周周报20260719.xlsx

    逻辑：YYYYMMDD 为提交日期（通常为周日），周报覆盖该日所在周（周一~周日）。
    """
    stem = os.path.splitext(os.path.basename(file_path))[0]
    # 匹配 YYYYMMDD 8位数字
    m = re.search(r'(\d{4})(\d{2})(\d{2})', stem)
    if m:
        try:
            d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            # 提交日期视为周报结束日（周日），周报周期为前6天到该日
            sunday = d
            monday = d - timedelta(days=6)
            return monday, sunday
        except ValueError:
            pass
    return None, None


def get_current_week(today: date = None) -> tuple[date, date]:
    if today is None:
        today = bj_today()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def classify_report_week(
    week_start: Optional[date], week_end: Optional[date]
) -> dict:
    today = bj_today()
    current_monday, current_sunday = get_current_week(today)

    result = {
        "report_type": "normal",
        "week_diff": 0,
        "needs_confirmation": False,
        "is_future": False,
        "message": "",
    }

    if week_start is None or week_end is None:
        result["report_type"] = "unknown"
        result["needs_confirmation"] = True
        result["message"] = "无法从文件中识别周报时间，请手动确认周报所属周次"
        return result

    if week_start > current_sunday:
        result["is_future"] = True
        result["report_type"] = "future"
        result["message"] = f"无法提交未来时间的周报。文件标注时间为 {week_start} ~ {week_end}，当前周为 {current_monday} ~ {current_sunday}"
        return result

    # 完全在当前周内
    if week_start >= current_monday and week_end <= current_sunday:
        result["report_type"] = "normal"
        result["week_diff"] = 0
        result["message"] = "本周周报"
        return result

    # 完全在过去（结束日期在当前周之前）
    if week_end < current_monday:
        diff_days = (current_monday - week_start).days
        weeks_behind = max(1, diff_days // 7)
        result["report_type"] = "catch_up"
        result["week_diff"] = weeks_behind
        result["message"] = f"补周报：这是 {weeks_behind} 周前的周报（{week_start} ~ {week_end}）"
        return result

    # 跨周情况：开始日期在当前周之前，但结束日期在当前周内或之后
    # 这种情况应该要求用户确认，而不是默认归类为"本周"
    result["report_type"] = "unknown"
    result["needs_confirmation"] = True
    result["message"] = f"周报时间跨周（{week_start} ~ {week_end}），请手动确认所属周次"
    return result


def _dates_from_text(text: str) -> tuple[Optional[date], Optional[date]]:
    """从文本中提取日期范围。
    支持格式：
      - 2026.6.29-2026.7.3  （4位年份）
      - 26.6.29-26.7.3      （2位年份，自动补全为20xx）
      - 2026-6-29 ~ 2026-7-3
      - 2026/6/29-2026/7/3
    """
    all_dates = []

    # 4位年份
    for m in re.finditer(r'(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text):
        try:
            d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            all_dates.append(d)
        except ValueError:
            continue

    # 2位年份（仅当4位年份未匹配到时使用）
    if not all_dates:
        for m in re.finditer(r'(\d{2})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', text):
            try:
                year = int(m.group(1))
                year += 2000 if year < 50 else 1900
                d = date(year, int(m.group(2)), int(m.group(3)))
                all_dates.append(d)
            except ValueError:
                continue

    if len(all_dates) >= 2:
        all_dates.sort()
        return all_dates[0], all_dates[-1]
    return None, None


# ── Excel 解析 ──

def _parse_excel(file_path: str) -> dict:
    wb = _safe_load_workbook(file_path)
    ws = wb.active

    result = {
        "title": ws.title,
        "last_week_work": [],
        "this_week_plan": [],
        "raw_content": "",
        "metadata": {},
    }

    current_section = None
    headers = []
    header_rows_for_meta = []  # 收集 section 切换前的顶部行，用于元数据扫描

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        first_cell = row[0]
        row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]

        if first_cell.value is None:
            header_rows_for_meta.append(row_values)
            continue

        cell_value = str(first_cell.value).strip()

        is_section_header = ("上周工作" in cell_value) or ("本周工作" in cell_value)
        if is_section_header:
            # 先把积累的 header_rows_for_meta 扫描一遍 metadata
            if header_rows_for_meta:
                found = _scan_metadata_from_rows(header_rows_for_meta)
                for k, v in found.items():
                    result["metadata"].setdefault(k, v)
                header_rows_for_meta = []
            if "上周工作" in cell_value:
                current_section = "last_week"
            elif "本周工作" in cell_value:
                current_section = "this_week"
            headers = []
            continue

        if current_section is None:
            header_rows_for_meta.append(row_values)
            continue

        if cell_value == "序号" or (len(row_values) > 1 and row_values[1] in ["客户/项目", "项目"]):
            headers = row_values
            continue

        if not headers:
            continue

        try:
            int(cell_value)
        except ValueError:
            # 非数字序号行，也可能是元数据行（如汇报人信息混入 section）
            header_rows_for_meta.append(row_values)
            continue

        item = {}
        for i, h in enumerate(headers):
            if i < len(row_values):
                item[h] = row_values[i]

        if current_section == "last_week":
            result["last_week_work"].append(item)
        elif current_section == "this_week":
            result["this_week_plan"].append(item)

    # 最后再扫一遍积累的 header_rows_for_meta（可能文件最后没有 section 标记）
    if header_rows_for_meta:
        found = _scan_metadata_from_rows(header_rows_for_meta)
        for k, v in found.items():
            result["metadata"].setdefault(k, v)

    result["raw_content"] = _build_text_content(result)
    wb.close()
    return result


def _extract_dates_from_excel(file_path: str) -> tuple[Optional[date], Optional[date]]:
    """从 Excel 文件中提取周报日期范围。
    扫描前 20 行，查找包含"上周"或"本周"的标题行。
    优先使用"上周工作内容"中的日期（更可靠），其次使用"本周工作计划"中的日期。
    """
    wb = _safe_load_workbook(file_path)
    ws = wb.active

    last_week_dates = None
    this_week_dates = None

    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), max_col=ws.max_column, values_only=False):
        first_cell = row[0]
        if first_cell.value is None:
            continue
        cell_value = str(first_cell.value)

        if "上周" in cell_value and "工作" in cell_value:
            d1, d2 = _dates_from_text(cell_value)
            if d1 and d2:
                last_week_dates = (d1, d2)
        elif "本周" in cell_value and "工作" in cell_value:
            d1, d2 = _dates_from_text(cell_value)
            if d1 and d2:
                this_week_dates = (d1, d2)

    wb.close()

    # 优先使用"上周工作内容"的日期（因为周报是回顾上周）
    if last_week_dates:
        return last_week_dates
    if this_week_dates:
        return this_week_dates
    return None, None


def parse_multi_week_excel(file_path: str) -> list[dict]:
    """解析包含多周数据的Excel文件。
    
    返回：
        list[dict]: 每个元素包含一周的数据，格式为：
        {
            "week_start": date,
            "week_end": date,
            "last_week_work": [...],
            "this_week_plan": [...],
            "raw_content": "...",
            "metadata": {...}
        }
    """
    wb = _safe_load_workbook(file_path)
    ws = wb.active
    
    weeks_data = []
    current_week = None
    current_section = None
    headers = []
    metadata = {}
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        first_cell = row[0]
        row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
        
        if first_cell.value is None:
            continue
        
        cell_value = str(first_cell.value).strip()
        
        # 检测周次标题行（包含日期范围）
        is_week_header = ("上周工作" in cell_value or "本周工作" in cell_value) and ("." in cell_value or "-" in cell_value or "/" in cell_value)
        
        if is_week_header:
            # 提取日期范围
            week_dates = _dates_from_text(cell_value)
            if not week_dates[0] or not week_dates[1]:
                continue
            
            # 如果是"上周工作内容"，开始新的一周数据
            if "上周工作" in cell_value:
                # 保存上一周数据（如果有）
                if current_week and (current_week["last_week_work"] or current_week["this_week_plan"]):
                    current_week["raw_content"] = _build_text_content(current_week)
                    current_week["metadata"] = metadata.copy()
                    weeks_data.append(current_week)
                
                # 开始新的一周
                current_week = {
                    "week_start": week_dates[0],
                    "week_end": week_dates[1],
                    "last_week_work": [],
                    "this_week_plan": [],
                }
                current_section = "last_week"
                headers = []
            elif "本周工作" in cell_value and current_week:
                # 切换到本周计划
                current_section = "this_week"
                headers = []
            continue
        
        if current_week is None:
            # 在第一个周次标题前的行可能是元数据
            if "汇报人" in cell_value or "部门" in cell_value:
                meta = _scan_metadata_from_rows([row_values])
                metadata.update(meta)
            continue
        
        # 检测表头行
        if cell_value == "序号" or (len(row_values) > 1 and row_values[1] in ["客户/项目", "项目"]):
            headers = row_values
            continue
        
        if not headers:
            continue
        
        # 解析数据行
        try:
            int(cell_value)
        except ValueError:
            continue
        
        item = {}
        for i, h in enumerate(headers):
            if i < len(row_values):
                item[h] = row_values[i]
        
        if current_section == "last_week":
            current_week["last_week_work"].append(item)
        elif current_section == "this_week":
            current_week["this_week_plan"].append(item)
    
    # 保存最后一周数据
    if current_week and (current_week["last_week_work"] or current_week["this_week_plan"]):
        current_week["raw_content"] = _build_text_content(current_week)
        current_week["metadata"] = metadata.copy()
        weeks_data.append(current_week)
    
    wb.close()
    return weeks_data


def extract_week_dates_from_content(content: str) -> tuple[Optional[date], Optional[date]]:
    """从周报文本内容中提取日期范围。
    支持格式：
      - "上周工作内容（2026.7.6-2026.7.12）"
      - "本周工作计划（2026.7.13-2026.7.17）"
      - "2026.7.6-2026.7.12"
      - "2026-7-6 ~ 2026-7-12"
    """
    # 尝试从内容中提取日期
    d1, d2 = _dates_from_text(content)
    if d1 and d2:
        return d1, d2
    return None, None


# ── Word / PDF 解析（已停用，当前业务仅支持 .xlsx）──
# 原 _parse_docx / _extract_dates_from_docx / _parse_pdf / _extract_dates_from_pdf
# 已于本轮改造中移除。仅保留 Excel 解析路径以符合"只允许 xlsx"的上传策略。





def _scan_metadata_from_rows(rows: list[list[str]]) -> dict[str, str]:
    """
    扫描二维表格形式的元数据。每行形如 ['汇报人', '张三'] 或 ['汇报人：张三']。
    （Excel row_values 会按 max_column 拉齐成多列，末尾为空串，所以这里会用 first_non_empty 处理）
    返回 {'reporter': '张三', 'department': '研发部'} 等。
    """
    meta = {}
    for row in rows:
        if not row:
            continue
        # 去掉尾部空串，获得真实有效列
        compact = [str(c or "").strip() for c in row]
        # 去掉尾部空值，获得真实结构
        while compact and not compact[-1]:
            compact.pop()
        if not compact:
            continue

        # 情况 A: 单格形式 "汇报人：张三" / "汇报人:张三"
        if len(compact) == 1:
            text = compact[0]
            for key, labels in _META_LABELS.items():
                if key in meta:
                    continue
                for label in labels:
                    for sep in ["：", ":", "\t"]:
                        prefix = label + sep
                        if text.startswith(prefix):
                            val = text[len(prefix):].strip()
                            if val:
                                meta[key] = val
                            break
                    if key in meta:
                        break
            continue

        # 情况 B: 多列形式，形如 ['汇报人', '张三', ...] 或 ['汇报人：张三', '', '']
        first_cell = compact[0]
        second_cell = compact[1] if len(compact) > 1 else ""

        # B-1: 第一格本身就是 "汇报人：张三" 形式
        for key, labels in _META_LABELS.items():
            if key in meta:
                continue
            for label in labels:
                for sep in ["：", ":", "\t"]:
                    prefix = label + sep
                    if first_cell.startswith(prefix):
                        val = first_cell[len(prefix):].strip()
                        if val:
                            meta[key] = val
                        break
                if key in meta:
                    break

        # B-2: 两列形式 "汇报人 | 张三"（当第一列本身就是标签、不带冒号）
        if first_cell and not first_cell.endswith("：") and not first_cell.endswith(":"):
            for key, labels in _META_LABELS.items():
                if key in meta:
                    continue
                for label in labels:
                    if first_cell == label or first_cell.rstrip("：:").strip() == label:
                        if second_cell:
                            meta[key] = second_cell
                        break

        # B-3: 两列表单冒号形式 "汇报人： | 张三"（A列标签带冒号，B列填值）
        if first_cell and (first_cell.endswith("：") or first_cell.endswith(":")):
            for key, labels in _META_LABELS.items():
                if key in meta:
                    continue
                for label in labels:
                    if first_cell.rstrip("：:").strip() == label:
                        if second_cell:
                            meta[key] = second_cell
                        break
    return meta


def _scan_metadata_from_flat_text(lines: list[str]) -> dict[str, str]:
    """
    从一整段非结构化文本中提取元数据。适用于 Word / PDF 段落。
    """
    meta = {}
    for line in lines:
        text = line.strip()
        if not text:
            continue
        for key, labels in _META_LABELS.items():
            if key in meta:
                continue
            for label in labels:
                for sep in ["：", ":", "\t"]:
                    prefix = label + sep
                    if text.startswith(prefix):
                        val = text[len(prefix):].strip()
                        if val:
                            meta[key] = val
                        break
                if key in meta:
                    break
    return meta


# ── 通用工具 ──

def _build_text_content(parsed: dict) -> str:
    lines = []

    if parsed["last_week_work"]:
        lines.append("## 上周工作内容")
        lines.append("")
        for i, item in enumerate(parsed["last_week_work"], 1):
            project = item.get("客户/项目", item.get("项目", ""))
            content = item.get("工作内容", "")
            reporter = item.get("汇报人", "")
            feedback = item.get("结果反馈", "")
            lines.append(f"### {i}. {project}")
            if content:
                lines.append(f"- 工作内容：{content}")
            if reporter:
                lines.append(f"- 汇报人：{reporter}")
            if feedback:
                lines.append(f"- 结果反馈：{feedback}")
            lines.append("")

    if parsed["this_week_plan"]:
        lines.append("## 本周工作计划")
        lines.append("")
        for i, item in enumerate(parsed["this_week_plan"], 1):
            project = item.get("客户/项目", item.get("项目", ""))
            content = item.get("工作内容", "")
            reporter = item.get("汇报人", "")
            feedback = item.get("结果反馈", "")
            lines.append(f"### {i}. {project}")
            if content:
                lines.append(f"- 工作内容：{content}")
            if reporter:
                lines.append(f"- 汇报人：{reporter}")
            if feedback:
                lines.append(f"- 预期产出：{feedback}")
            lines.append("")

    return "\n".join(lines)
