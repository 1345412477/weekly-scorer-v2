"""企业微信数据解析：考勤打卡 Excel + 聊天记录 Excel."""
import re
import logging
from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Tuple, Optional

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

logger = logging.getLogger(__name__)


# -------- 通用工具 --------

def _norm(text: Any) -> str:
    """归一化单元格文本：去除空白、None 安全"""
    if text is None:
        return ""
    return str(text).strip()


def _parse_date(value: Any) -> Optional[date]:
    """从多种格式（date/datetime/字符串）解析日期"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _norm(value)
    # 常见中文/标准格式
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # 兼容 openpyxl 的「2026/6/1 0:00:00」
    m = re.search(r"(\d{4})[-/\.年](\d{1,2})[-/\.月](\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _parse_datetime(value: Any) -> Optional[datetime]:
    """解析完整日期时间（用于聊天记录发送时间）"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = _norm(value)
    # 常见完整格式
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # 先尝试日期部分
    d = _parse_date(s)
    if d is not None:
        # 尝试抽时间
        m = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
        if m:
            h = int(m.group(1))
            mi = int(m.group(2))
            se = int(m.group(3)) if m.group(3) else 0
            return datetime(d.year, d.month, d.day, h, mi, se)
        return datetime(d.year, d.month, d.day)
    return None


def _parse_time(value: Any) -> Optional[str]:
    """解析打卡时间为 HH:MM 字符串，未打卡返回 None"""
    if value is None or value == "":
        return None
    s = _norm(value)
    # 直接 HH:MM / HH:MM:SS
    m = re.search(r"(\d{1,2}):(\d{2})(?::\d{2})?", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    return None


def _week_range(target: date) -> Tuple[date, date]:
    monday = target - timedelta(days=target.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _strip_wecom_role(text: str) -> str:
    """剥离企业微信标识后缀，例如「曾静@员工」→「曾静」，「研发部@群聊」→「研发部」"""
    t = _norm(text)
    if "@" in t:
        return t.split("@", 1)[0].strip()
    return t


# -------- 考勤解析 --------

# 可能的字段名映射（企业微信不同版本/语言有差异）
_ATTENDANCE_FIELD_MAP = {
    "姓名": "author_name",
    "成员": "author_name",
    "员工": "author_name",
    "日期": "record_date",
    "时间": "record_date",
    # "打卡时间" 移除，避免在两行表头中误匹配（二级表头"打卡时间"不应被覆盖）
    "打卡日期": "record_date",
    "部门": "department",
    "所属部门": "department",
    "打卡类型": "punch_type",
    "类型": "punch_type",
    "应打卡时间": "should_time",
    "实际打卡时间": "actual_time",
    "打卡时间": "actual_time",
    "打卡地点": "location",
    "地点": "location",
    "打卡状态": "status",
    "状态": "status",
    "备注内容": "notes",
    "备注": "notes",
    "说明": "notes",
    "假勤申请": "leave_request",
    # 新版企业微信导出字段
    "最早": "check_in_time",
    "最晚": "check_out_time",
    "考勤结果": "status",
    "上班打卡时间": "check_in_time",
    "下班打卡时间": "check_out_time",
    "迟到次数": "late_count",
    "早退次数": "early_count",
    "缺卡次数": "miss_count",
    "补卡次数": "supplement_count",
    "实际工作时长(小时)": "work_hours",
    "加班时长(小时)": "overtime_hours",
}

# 新版两行表头组合映射：(一级表头, 二级表头) -> field
_ATTENDANCE_COMBINED_MAP = {
    ("上班1", "打卡时间"): "check_in_time",
    ("下班1", "打卡时间"): "check_out_time",
    ("上班1", "打卡状态"): "check_in_status",
    ("下班1", "打卡状态"): "check_out_status",
    # "时间" 列作为 record_date（仅当无二级表头时）
    ("时间", ""): "record_date",
}

# 通配符组合：任意一级表头 + 特定二级表头 -> field
_ATTENDANCE_SECONDARY_ONLY = {
    "最早": "check_in_time",
    "最晚": "check_out_time",
    "班次": "shift_type",
    "考勤结果": "status",
    "打卡时间": "actual_time",
    "打卡状态": "status",
}


def _match_att_field(header: str) -> Optional[str]:
    header = _norm(header)
    if not header:
        return None
    # 精确匹配优先
    if header in _ATTENDANCE_FIELD_MAP:
        return _ATTENDANCE_FIELD_MAP[header]
    # 子串匹配：要求 header 以 key 开头或结尾，避免 "打卡时间记录" 误匹配 "打卡日期"
    for key, field in _ATTENDANCE_FIELD_MAP.items():
        if header.startswith(key) or header.endswith(key):
            return field
    return None


def _find_attendance_sheet(wb) -> str:
    """优先选择「打卡详情」sheet，若无则选第一个 sheet"""
    # 精确匹配优先
    exact = ["打卡详情", "打卡明细"]
    for name in wb.sheetnames:
        if name in exact:
            return name
    # 模糊匹配：包含"打卡"或"考勤"的 sheet
    # 注意：新版企业微信导出 sheet 名可能是"概况统计与打卡明细"，需要接受
    for name in wb.sheetnames:
        if "打卡" in name or "考勤" in name:
            return name
    return wb.active.title


def _detect_two_row_header(rows) -> Tuple[bool, int]:
    """检测是否为新版两行表头格式。
    返回 (is_two_row, header_row_idx)。
    两行表头特征：某行的前几列包含「姓名」但无「日期」，下一行包含「最早/班次/考勤结果」等二级字段。
    只看前 10 列以避免后面"打卡详情"等干扰列的影响。
    """
    for idx in range(min(8, len(rows) - 1)):
        row_prefix = "".join(_norm(c) for c in rows[idx][:10])
        next_text = "".join(_norm(c) for c in rows[idx + 1])
        # 一级表头前几列有「姓名」但无「日期」
        if ("姓名" in row_prefix or "员工" in row_prefix) and "日期" not in row_prefix:
            # 下一行有二级字段
            if "最早" in next_text or "班次" in next_text or "考勤结果" in next_text:
                return True, idx
    return False, -1


# 新版格式中作为分组标题、不应映射为数据字段的列名
_TWO_ROW_SKIP_HEADERS = {"打卡时间记录", "打卡详情", "基础信息", "考勤概况", "异常统计", "外出打卡", "加班统计", "假勤统计"}


def _build_two_row_field_cols(rows, header_row_idx: int) -> Dict[int, str]:
    """根据两行表头构建 col -> field 映射。
    一级表头在 header_row_idx，二级表头在 header_row_idx+1。
    """
    primary = [_norm(c) for c in rows[header_row_idx]]
    secondary = [_norm(c) for c in rows[header_row_idx + 1]]
    field_cols: Dict[int, str] = {}

    for col_idx in range(max(len(primary), len(secondary))):
        p = primary[col_idx] if col_idx < len(primary) else ""
        s = secondary[col_idx] if col_idx < len(secondary) else ""

        # 跳过分组标题列
        if p in _TWO_ROW_SKIP_HEADERS:
            continue

        # 优先用组合映射
        if p and s and (p, s) in _ATTENDANCE_COMBINED_MAP:
            field_cols[col_idx] = _ATTENDANCE_COMBINED_MAP[(p, s)]
            continue

        # 一级表头单独匹配
        if p and not s:
            field = _match_att_field(p)
            if field:
                field_cols[col_idx] = field
                continue

        # 二级表头单独匹配
        if s:
            field = _match_att_field(s)
            if field:
                field_cols[col_idx] = field
                continue

    return field_cols


def parse_attendance_excel(file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    解析企业微信打卡 Excel。
    兼容两种格式：
    - 旧版：单行表头，每行一条打卡记录（含打卡类型区分上下班）
    - 新版：两行表头，每行一个员工一天的汇总（含上班1/下班1打卡时间列）
    返回 (records, unmatched_names)。
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl 未安装，无法解析 Excel")

    wb = load_workbook(file_path, data_only=True)
    ws_name = _find_attendance_sheet(wb)
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []

    # 检测表头格式
    is_two_row, header_row_idx = _detect_two_row_header(rows)

    if is_two_row:
        return _parse_new_format(rows, header_row_idx)
    else:
        # 单行表头：在前 10 行中找到包含「姓名」或「员工」的行作为表头
        header_row_idx = _find_single_row_header(rows)
        if header_row_idx < 0:
            logger.warning(f"[考勤解析] 未找到表头行，文件: {file_path}")
            return [], []
        return _parse_old_format(rows, header_row_idx)


def _find_single_row_header(rows) -> int:
    """在单行表头模式下找到表头行索引。
    在前 10 行中查找包含「姓名」或「员工」的行。
    """
    for idx in range(min(10, len(rows))):
        row_text = "".join(_norm(c) for c in rows[idx])
        if "姓名" in row_text or "员工" in row_text or "成员" in row_text:
            return idx
    return -1


def _parse_old_format(rows, header_row_idx: int) -> Tuple[List[Dict[str, Any]], List[str]]:
    """旧版单行表头解析（原有逻辑）"""
    headers = [_norm(c) for c in rows[header_row_idx]]
    field_cols: Dict[int, str] = {}
    for col_idx, h in enumerate(headers):
        field = _match_att_field(h)
        if field:
            field_cols[col_idx] = field

    buckets: Dict[Tuple[str, date], Dict[str, Any]] = {}

    for row in rows[header_row_idx + 1:]:
        if all(_norm(c) == "" for c in row):
            continue
        raw: Dict[str, Any] = {}
        for col_idx, field_name in field_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if field_name == "record_date":
                raw["record_date"] = _parse_date(val)
            elif field_name == "should_time":
                raw["should_time"] = _parse_time(val)
            elif field_name == "actual_time":
                raw["actual_time"] = _parse_time(val)
            else:
                raw[field_name] = _norm(val) if val is not None else ""

        if not raw.get("author_name"):
            continue
        if not raw.get("record_date"):
            continue

        author = raw["author_name"]
        rdate = raw["record_date"]
        ptype = raw.get("punch_type", "") or ""
        atime = raw.get("actual_time")
        loc = raw.get("location", "") or ""
        dept = raw.get("department", "") or ""
        status = raw.get("status", "") or ""
        notes = raw.get("notes", "") or ""
        leave = raw.get("leave_request", "") or ""

        key = (author, rdate)
        if key not in buckets:
            ws_date, we_date = _week_range(rdate)
            buckets[key] = {
                "author_name": author,
                "department": dept,
                "record_date": rdate,
                "week_start": ws_date,
                "week_end": we_date,
                "check_in_time": None,
                "check_out_time": None,
                "check_in_location": None,
                "check_out_location": None,
                "attendance_status": "",
                "notes": "",
            }

        bucket = buckets[key]

        status_parts = []
        if status and status not in bucket["attendance_status"]:
            status_parts.append(status)
        if leave and leave not in bucket["attendance_status"]:
            status_parts.append(leave)
        if status_parts:
            prev = bucket["attendance_status"]
            bucket["attendance_status"] = (prev + " / " + " / ".join(status_parts)).strip(" / ")
        if notes and notes not in bucket["notes"]:
            prev = bucket["notes"]
            bucket["notes"] = (prev + " / " + notes).strip(" / ")

        if "上班" in ptype and atime:
            if bucket["check_in_time"] is None or atime < bucket["check_in_time"]:
                bucket["check_in_time"] = atime
                bucket["check_in_location"] = loc or bucket.get("check_in_location")
        elif "下班" in ptype and atime:
            if bucket["check_out_time"] is None or atime > bucket["check_out_time"]:
                bucket["check_out_time"] = atime
                bucket["check_out_location"] = loc or bucket.get("check_out_location")
        elif ("外出" in ptype or "打卡" in ptype) and atime:
            if bucket["check_in_time"] is None:
                bucket["check_in_time"] = atime
                bucket["check_in_location"] = loc
            elif bucket["check_out_time"] is None and atime > bucket["check_in_time"]:
                bucket["check_out_time"] = atime
                bucket["check_out_location"] = loc

    return _finalize_buckets(buckets)


def _parse_new_format(rows, header_row_idx: int) -> Tuple[List[Dict[str, Any]], List[str]]:
    """新版两行表头解析。
    每行是一个员工一天的汇总，直接包含上班/下班打卡时间列。
    表头结构：
      Row N:   时间 | 姓名 | 账号 | 基础信息 | ... | 上班1 | ... | 下班1 | ...
      Row N+1:      |      |      | 部门|职务|工号 | ... | 打卡时间|打卡状态 | 打卡时间|打卡状态 | ...
    """
    field_cols = _build_two_row_field_cols(rows, header_row_idx)
    data_start = header_row_idx + 2  # 跳过两行表头

    buckets: Dict[Tuple[str, date], Dict[str, Any]] = {}

    for row in rows[data_start:]:
        if all(_norm(c) == "" for c in row):
            continue

        raw: Dict[str, Any] = {}
        for col_idx, field_name in field_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if field_name == "record_date":
                raw["record_date"] = _parse_date(val)
            elif field_name in ("check_in_time", "check_out_time"):
                raw[field_name] = _parse_time(val)
            elif field_name == "should_time":
                raw["should_time"] = _parse_time(val)
            else:
                raw[field_name] = _norm(val) if val is not None else ""

        author = raw.get("author_name", "")
        if not author:
            continue
        rdate = raw.get("record_date")
        if not rdate:
            continue

        key = (author, rdate)
        if key not in buckets:
            ws_date, we_date = _week_range(rdate)
            buckets[key] = {
                "author_name": author,
                "department": raw.get("department", ""),
                "record_date": rdate,
                "week_start": ws_date,
                "week_end": we_date,
                "check_in_time": None,
                "check_out_time": None,
                "check_in_location": None,
                "check_out_location": None,
                "attendance_status": "",
                "notes": "",
            }

        bucket = buckets[key]

        # 新版格式直接有 check_in_time / check_out_time
        if raw.get("check_in_time"):
            bucket["check_in_time"] = raw["check_in_time"]
        if raw.get("check_out_time"):
            bucket["check_out_time"] = raw["check_out_time"]

        # 考勤结果作为状态
        status = raw.get("status", "") or ""
        if status and status != "--" and status not in bucket["attendance_status"]:
            prev = bucket["attendance_status"]
            bucket["attendance_status"] = (prev + " / " + status).strip(" / ")

    return _finalize_buckets(buckets)


def _finalize_buckets(buckets: Dict) -> Tuple[List[Dict[str, Any]], List[str]]:
    """计算工时、排序、返回结果"""
    records: List[Dict[str, Any]] = []
    for bucket in buckets.values():
        if bucket["check_in_time"] and bucket["check_out_time"]:
            try:
                h1, m1 = bucket["check_in_time"].split(":")
                h2, m2 = bucket["check_out_time"].split(":")
                diff_h = (int(h2) * 60 + int(m2) - int(h1) * 60 - int(m1)) / 60.0
                if diff_h > 0:
                    bucket["work_duration_hours"] = round(diff_h, 2)
                else:
                    bucket["work_duration_hours"] = None
            except (ValueError, AttributeError):
                bucket["work_duration_hours"] = None
        else:
            bucket["work_duration_hours"] = None
        records.append(bucket)

    records.sort(key=lambda r: (r["record_date"], r["author_name"]))
    unmatched = sorted({r["author_name"] for r in records})
    logger.info(f"[考勤解析] 读取 {len(records)} 条员工-日记录，涉及 {len(unmatched)} 位员工")
    return records, unmatched


def summarize_attendance_for_person(
    records: List[Dict[str, Any]], author_name: str
) -> str:
    """将某员工本周考勤记录转为 AI 评分摘要文本"""
    filtered = [r for r in records if r.get("author_name") == author_name]
    if not filtered:
        return f"{author_name}: 本周无考勤数据"

    lines = [f"员工 {author_name} 本周 {len(filtered)} 条打卡记录："]
    for r in filtered:
        parts = []
        parts.append(f"日期 {r.get('record_date', '-')}")
        parts.append(f"上班 {r.get('check_in_time', '未打卡')}@{r.get('check_in_location', '无地点')}")
        parts.append(f"下班 {r.get('check_out_time', '未打卡')}@{r.get('check_out_location', '无地点')}")
        dur = r.get("work_duration_hours")
        if dur is not None:
            parts.append(f"工时 {dur}h")
        status = r.get("attendance_status")
        if status:
            parts.append(f"状态 {status}")
        notes = r.get("notes")
        if notes:
            parts.append(f"备注 {notes}")
        lines.append("  · " + " ｜ ".join(parts))
    return "\n".join(lines)


# -------- 聊天记录解析 --------

_CHAT_FIELD_MAP = {
    "分组": "group_name",
    "发送者": "sender",
    "接收者": "receiver",
    "内容": "content",
    "会话类型": "session_type",
    "消息类型": "message_type",
    "发送时间": "send_time",
    # 兼容旧版字段名（若有）
    "姓名": "sender",
    "员工": "sender",
    "日期": "send_time",
    "聊天日期": "send_time",
    "会话主题": "group_name",
    "主题": "group_name",
    "对方": "receiver",
    "聊天对象": "receiver",
    "消息数": "message_count",
    "数量": "message_count",
    "响应时长": "response_minutes",
    "平均响应": "response_minutes",
    "内容摘要": "content",
    "摘要": "content",
}


def _match_chat_field(header: str) -> Optional[str]:
    header = _norm(header)
    if not header:
        return None
    if header in _CHAT_FIELD_MAP:
        return _CHAT_FIELD_MAP[header]
    for key, field in _CHAT_FIELD_MAP.items():
        if key in header:
            return field
    return None


def _extract_chat_dates_from_filename(file_path: str) -> Tuple[Optional[date], Optional[date]]:
    """从聊天记录文件名中提取日期范围，作为发送时间的 fallback。
    支持格式：20260706-20260712、20260706_20260712、2026-07-06_2026-07-12
    返回 (start_date, end_date)。
    """
    fname = os.path.basename(file_path)
    m = re.search(r"(\d{4})[-_]?(\d{1,2})[-_]?(\d{1,2})[-_](\d{4})[-_]?(\d{1,2})[-_]?(\d{1,2})", fname)
    if m:
        try:
            start = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            end = date(int(m.group(4)), int(m.group(5)), int(m.group(6)))
            return start, end
        except ValueError:
            pass
    return None, None


def parse_chat_excel(file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    解析企业微信聊天记录 Excel。
    原始数据为逐条消息，解析后按（员工, 周）聚合并计算 message_count 与粗略平均响应时长。
    返回 (records, unmatched_names)。
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl 未安装，无法解析 Excel")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []

    # 定位表头行
    header_row_idx = 0
    for idx, row in enumerate(rows[:5]):
        text = "".join(_norm(c) for c in row)
        if "发送者" in text or ("姓名" in text and "发送" in text) or ("内容" in text and "时间" in text):
            header_row_idx = idx
            break

    headers = [_norm(c) for c in rows[header_row_idx]]
    field_cols: Dict[int, str] = {}
    for col_idx, h in enumerate(headers):
        field = _match_chat_field(h)
        if field:
            field_cols[col_idx] = field

    has_send_time = "send_time" in field_cols.values()
    # 无发送时间列时，从文件名推算日期作为 fallback
    fn_start, fn_end = (None, None)
    if not has_send_time:
        fn_start, fn_end = _extract_chat_dates_from_filename(file_path)

    # 原始消息
    raw_messages: List[Dict[str, Any]] = []
    for row in rows[header_row_idx + 1:]:
        if all(_norm(c) == "" for c in row):
            continue
        msg: Dict[str, Any] = {}
        for col_idx, field_name in field_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if field_name == "send_time":
                msg["send_time"] = _parse_datetime(val)
            else:
                msg[field_name] = _norm(val) if val is not None else ""

        sender_raw = msg.get("sender", "")
        sender = _strip_wecom_role(sender_raw)
        if not sender:
            continue
        msg["sender"] = sender

        receiver_raw = msg.get("receiver", "")
        msg["receiver"] = _strip_wecom_role(receiver_raw)

        # 无发送时间时，用文件名日期作为 fallback
        if not has_send_time and msg.get("send_time") is None and fn_start:
            msg["send_time"] = datetime(fn_start.year, fn_start.month, fn_start.day, 12, 0, 0)

        raw_messages.append(msg)

    # 按（sender, week）聚合消息
    # 同时按（sender, group_name, day）估算响应时长
    bucket_key: Dict[Tuple[str, date, date], Dict[str, Any]] = {}

    # 用于计算响应时长：按（sender, group_name, day）分组后排序
    # 但响应时长通常指"对方回复的时间差"，这里用"同分组内连续消息间隔"的均值作为粗略指标
    # 先按 group 聚合消息序列
    by_group: Dict[Tuple[str, str], List[datetime]] = {}
    for msg in raw_messages:
        sender = msg["sender"]
        group = msg.get("group_name") or msg.get("receiver") or ""
        send_time = msg.get("send_time")
        if send_time is None:
            continue
        key = (sender, group)
        if key not in by_group:
            by_group[key] = []
        by_group[key].append(send_time)

    # 计算每人响应时长均值（分钟）
    response_per_sender: Dict[str, List[float]] = {}
    for (sender, _group), times in by_group.items():
        if len(times) < 2:
            continue
        times_sorted = sorted(times)
        for i in range(1, len(times_sorted)):
            diff = (times_sorted[i] - times_sorted[i - 1]).total_seconds() / 60.0
            # 只保留 0-720 分钟的合理间隔（避免跨天异常大值）
            if 0 < diff <= 720:
                response_per_sender.setdefault(sender, []).append(diff)

    # 按周聚合
    for msg in raw_messages:
        sender = msg["sender"]
        send_time = msg.get("send_time")
        if send_time is None:
            continue
        msg_date = send_time.date()
        ws_date, we_date = _week_range(msg_date)
        key = (sender, ws_date, we_date)

        if key not in bucket_key:
            bucket_key[key] = {
                "author_name": sender,
                "week_start": ws_date,
                "week_end": we_date,
                "message_date": msg_date,
                "message_count": 0,
                "conversation_topic": "",
                "counterparty": "",
                "content_summary": "",
                "topics_set": set(),
                "counterparty_set": set(),
                "snippets": [],
            }

        b = bucket_key[key]
        b["message_count"] += 1
        group = msg.get("group_name") or ""
        receiver = msg.get("receiver") or ""
        if group:
            b["topics_set"].add(group)
        if receiver:
            b["counterparty_set"].add(receiver)
        content = msg.get("content") or ""
        mtype = msg.get("message_type") or ""
        if content and mtype != "撤回了一条消息":
            snippet = content[:80]
            if snippet and snippet not in b["snippets"]:
                b["snippets"].append(snippet)

    # 整理为最终记录
    records: List[Dict[str, Any]] = []
    for b in bucket_key.values():
        # 填充响应时长
        sender = b["author_name"]
        resp_list = response_per_sender.get(sender, [])
        resp_avg = round(sum(resp_list) / len(resp_list), 2) if resp_list else None

        topics_list = list(b["topics_set"])
        counterparty_list = list(b["counterparty_set"])

        summary_parts = []
        if b["snippets"]:
            summary_parts.append("; ".join(b["snippets"][:3]))

        records.append({
            "author_name": sender,
            "week_start": b["week_start"],
            "week_end": b["week_end"],
            "message_date": b["message_date"],
            "conversation_topic": ", ".join(topics_list[:3]),
            "counterparty": ", ".join(counterparty_list[:5]),
            "message_count": b["message_count"],
            "response_minutes": resp_avg,
            "content_summary": " | ".join(summary_parts)[:500],
        })

    records.sort(key=lambda r: (r["week_start"], r["author_name"]))
    unmatched = sorted({r["author_name"] for r in records})
    logger.info(f"[聊天记录解析] 读取 {len(records)} 条记录，涉及 {len(unmatched)} 位员工")
    return records, unmatched


def summarize_chat_for_person(
    records: List[Dict[str, Any]], author_name: str, weekly_summaries: Optional[List[Dict[str, Any]]] = None
) -> str:
    """将某员工本周聊天+一周小结合并为 AI 评分摘要文本"""
    filtered = [r for r in records if r.get("author_name") == author_name]
    summaries = [s for s in (weekly_summaries or []) if s.get("author_name") == author_name]

    lines = [f"员工 {author_name} 本周沟通数据："]
    if filtered:
        total = sum(r.get("message_count", 0) for r in filtered)
        lines.append(f"聊天记录 {len(filtered)} 组，共 {total} 条消息：")
        for r in filtered:
            parts = [
                f"日期范围 {r.get('week_start', '-')} ~ {r.get('week_end', '-')}",
                f"主题 {r.get('conversation_topic', '无')}",
                f"对方 {r.get('counterparty', '无')}",
                f"消息数 {r.get('message_count', 0)}",
            ]
            resp = r.get("response_minutes")
            if resp is not None:
                parts.append(f"平均响应 {resp} 分钟")
            summary = r.get("content_summary")
            if summary:
                parts.append(f"摘要：{summary}")
            lines.append("  · " + " ｜ ".join(parts))
    else:
        lines.append("  （无聊天记录）")

    if summaries:
        lines.append(f"一周小结 {len(summaries)} 条：")
        for s in summaries:
            parts = [
                f"工作会话 {s.get('work_session_count', '未知')} 次",
                f"总耗时 {s.get('total_minutes', '未知')} 分钟",
                f"最晚 {s.get('latest_time', '未知')}",
            ]
            lines.append("  · " + " ｜ ".join(parts))
    else:
        lines.append("  （无一周小结）")

    return "\n".join(lines)
